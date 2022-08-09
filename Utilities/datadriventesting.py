import softest as softest
from openpyxl import load_workbook



class Utils(softest.TestCase):

    def read_data_from_excel(file_name,sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct+1):
                row.append(sh.cell(row=1,column=1).value)
            datalist.append(row)
        return datalist


